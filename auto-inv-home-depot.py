import openpyxl
from PIL import Image
from bs4 import BeautifulSoup
import urllib
import urllib.request
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from time import sleep
from requests_html import HTMLSession

def get_info(url):
    options = Options()
    options.headless = True
    options.add_argument("--window-size=1920,1200")
    
    DRIVER_PATH = 'chromedriver'
    driver = webdriver.Chrome(options=options, executable_path=DRIVER_PATH)
    driver.get(url)
    sleep(2)
    html = driver.execute_script("return document.getElementsByTagName('html')[0].innerHTML")
    soup = BeautifulSoup(html, features="html.parser")
    driver.quit()
    session = HTMLSession()
    r = session.get(url)
    #retrieve image
    if soup.find(id="thd-helmet__link--preloadImg") == None:
        if soup.find("div", {"class":"product-identifier product-identifier__model"}) != None:
            url = soup.find("div", {"class":"product-identifier product-identifier__model"}).get_text().split(" ")[1]
            print("\n\n\n\n\nFound a list of different items would you like to continue the script with an item with the following model number:",url,"?\n")
            get_info("https://www.homedepot.com/s/{}?".format( url ))
            return
        print("\n\n\n\n\nCould not find item, sorry!\n\n")
        sleep(5)
        return
    img_link_tag = soup.find(id="thd-helmet__link--preloadImg")
    img_link = img_link_tag.get('href')
    sm_img_link="./pictures/{}".format( img_link.split('/')[-1] )
    urllib.request.urlretrieve(img_link, sm_img_link)
    img1 = Image.open(sm_img_link,mode='r')
    img2 = Image.open("logo.jpg",mode='r')
    img1.paste(img2)
    img1.save(sm_img_link)
    
    wb = openpyxl.load_workbook("test.xlsx")
    ws = wb.active
    i = str(ws.max_row + 1)
    #pallet
    ifUsed = input("Is this used? Enter 'y' for used or 'n' for new. \n")
    if ifUsed == 'y':
        ws['C' + i] = "used"
    elif ifUsed == 'n':
        ws['C' + i] = "new"
    #name
    ws['B' + i] = soup.find("title", {"data-th":"server"}).get_text()
    #box status
    ifBox = input("Is the item in box? Answer 'y' for yes and 'n' for no.\n")
    if ifBox == 'y':
        ws['D' + i] = "in box"
    elif ifBox == 'n':
        ws['D' + i] = "no box"
    #upc
    if  r.html.search('"upc":"{}"') != None:
        ws['E' + i] = r.html.search('"upc":"{}"')[0]
    #sku
    if r.html.search('"storeSkuNumber":"{}"') != None:
        ws['F' + i] = r.html.search('"storeSkuNumber":"{}"')[0]
    #price
    pricetag = soup.find("div", {"class":"price"})
    price = pricetag.get_text()
    ws['G' + i] = "{}.{}".format(price[1:-2], price[-2]+price[-1])
#old price
    if soup.find("span", {"class":"u__strike"}) != None:
        price = soup.find("span", {"class":"u__strike"}).get_text()
        #ws['H' + i] = "{}.{}".format(price[1:-2], price[-2]+price[-1])
        ws['H' + i] = price
    wb.save("test.xlsx")

def main():
	stop = False
	while(stop == False):
		if input("Would you like to enter a url or item number? enter 'u' for URL and 'i' for item number\n") == 'i':
			url = input("Enter a UPC code or anything that should work in HomeDepot.com search bar\n")
			get_info("https://www.homedepot.com/s/{}?".format( url ))
		else:
			url = input("Enter a Homedepot URL\n")
			get_info(url)
		if input("Another item? y for yes, n for no\n") == "n":
        		stop = True

if __name__ == "__main__":
    main()
