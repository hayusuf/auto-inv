import openpyxl
from PIL import Image
from bs4 import BeautifulSoup
import urllib
import urllib.request
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from time import sleep
from requests_html import HTMLSession

def get_info(url):
    #options = Options()
    #options.headless = True
    #options.add_argument("--window-size=1920,1200")
    
    #DRIVER_PATH = 'chromedriver'
    #driver = webdriver.Chrome(options=options, executable_path=DRIVER_PATH)
    #driver.get(url)
    #sleep(5)
    #html = driver.execute_script("return document.getElementsByTagName('html')[0].innerHTML")
    #soup = BeautifulSoup(html, features="html.parser")
    #driver.quit()
    session = HTMLSession()
    r = session.get(url)
    #retrieve image 
    img_link = r.html.search('<meta itemprop="image" content="{}" />')[0]
    print(img_link)
    #img_link_tag = soup.find(id="thd-helmet__link--preloadImg")
    #img_link = img_link_tag.get('href')
    sm_img_link = r.html.search('<meta property="og:url" content="https://www.kohls.com/product/{}/{}.jsp" />')[1] + ".jpg"
    print(sm_img_link)
    urllib.request.urlretrieve(img_link, sm_img_link)
    img1 = Image.open(sm_img_link,mode='r')
    img2 = Image.open("logo.jpg",mode='r')
    img1.paste(img2)
    img1.save(sm_img_link)
    
    wb = openpyxl.load_workbook("test.xlsx")
    ws = wb.active
    i = str(ws.max_row + 1)
    print(i)
    #pallet
    ws['A' + i] = input("What pallet number is this item from?\n")
    #name
    ws['B' + i] = r.html.search('<meta itemprop="name" content="{}" />')[0]
    #box status
    #ifBox = input("Is the item in box? Answer 'y' for yes and 'n' for no.\n")
    #if ifBox == 'y':
    #    ws['C' + i] = "in box"
    #elif ifBox == 'n':
    #    ws['C' + i] = "no box"
    #upc
    #if  r.html.search('"upc":"{}"') != None:
        #ws['E' + i] = r.html.search('"upc":"{}"')[0]
    #sku
    #if r.html.search('"storeSkuNumber":"{}"') != None:
        #ws['F' + i] = r.html.search('"storeSkuNumber":"{}"')[0]
    #price
    ws['G' + i] = r.html.search('<meta itemprop="price" content="{}" />')[0]
    wb.save("test.xlsx")

def main():
    get_info(input("Enter kohls item url\n"))

if __name__ == "__main__":
    main()
