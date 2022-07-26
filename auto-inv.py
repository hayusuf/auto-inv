import openpyxl
from bs4 import BeautifulSoup
import urllib
import urllib.request
from os.path import exists

def make_soup(url):
    html = urllib.request.urlopen(url).read()
    return BeautifulSoup(html, features="html.parser")

def get_images(url, pallet):
    soup = make_soup(url)

    tag = soup.find(id="thd-helmet__link--preloadImg")
    if tag is None:
        print('No matching item and image found')
        return
    link = tag.get('href')
    if exists("{}_second".format( link.split('/')[-1] )):
       urllib.request.urlretrieve(link, "{}_third".format( link.split('/')[-1] ))    
    elif exists("{}".format( link.split('/')[-1] )):
       urllib.request.urlretrieve(link, "{}_second".format( link.split('/')[-1] ))    
    else:
        urllib.request.urlretrieve(link, "{}".format( link.split('/')[-1] ))
    return
    
def get_name(url):
    soup = make_soup(url)

    tag = soup.find("title", {"data-th":"server"})
    if tag is None:
        print('No matching name')
        return
    name = tag.get_text()
    return name

def get_SKU(url):
    soup = make_soup(url)

    tag = soup.find_all("h2", {"class":"product-info-bar__detail--7va8o"})[2]
    if tag is None:
        print('No matching SKU')
        return
    SKU = tag.get_text()
    return SKU[-7:-1]

def get_UPC(url):
    soup = make_soup(url)

    tag = soup.find_all("h2", {"class":"product-info-bar__detail--7va8o"})[2]
    if tag is None:
        print('No matching UPC')
        return
    UPC = tag.get_text()
    return UPC

def get_price(url):
    soup = make_soup(url)

    tag = soup.find("div", {"id":"standard-price"})
    if tag is None:
        print('No matching price')
        return
    price = tag.get_text()
    return "{}.{}".format(price[1:-2], price[-2]+price[-1])

def write(inp, url, pallet):
    wb = openpyxl.load_workbook("test.xlsx")
    ws = wb.active
    i = 1
    cell_val = ''
    while cell_val != '':
        cell_val = ws['A' + i].value
        i += 1
    i = str(i)
    ws['A' + i] = pallet
    ws['B' + i] = get_name(url)
    ifBox = input("Is the item in box? Answer 'y' for yes and 'n' for no.\n")
    if ifBox == 'y':
        ws['C' + i] = "in box"
    elif ifBox == 'n':
        ws['C' + i] = "no box"
    isUPC = input("Is the code you entered a UPC or SKU or Model or neither? Enter 'u' for UPC and 's' for SKU and 'm' for model and 'n' for neither.\n")
    if isUPC == 'u':
        ws['E' + i] = inp
        ws['F' + i] = get_SKU(url)
    elif isUPC == 's':
        ws['E' + i] = get_UPC(url)
        ws['F' + i] = inp
    elif isUPC == 'm':
        ws['E' + i] = get_UPC(url)
        ws['F' + i] = get_SKU(url)
    elif isUPC == 'n':
        ws['E' + i] = get_UPC(url)
        ws['F' + i] = get_SKU(url)
    ws['G' + i] = get_price(url)
    wb.save("test.xlsx")

def main():
    inp = input("Enter a UPC code or anything that should work in HomeDepot.com search bar\n")
    pallet = input("What pallet number is this item from(what folder should it be added to)?\n")
    get_images("https://www.homedepot.com/s/{}?".format( inp ), pallet)
    write(inp,"https://www.homedepot.com/s/{}?".format( inp ), pallet)

if __name__ == "__main__":
    main()
