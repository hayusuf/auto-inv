import openpyxl
from requests_html import HTMLSession
import urllib
import urllib.request
from os.path import exists

def get_info(url):
    session = HTMLSession()
    r = session.get(url)
    r.html.render(timeout=100)

    img_link = r.html.search('<img src="{}"')[0]
    if img_link is None:
        print('No matching item and image found')
    if exists("{}_second".format( img_link.split('/')[-1] )):
       urllib.request.urlretrieve(img_link , "{}_third".format( img_link.split('/')[-1] ))    
    elif exists("{}".format( img_link.split('/')[-1] )):
       urllib.request.urlretrieve(img_link, "{}_second".format( img_link.split('/')[-1] ))    
    else:
        urllib.request.urlretrieve(img_link, "{}".format( img_link.split('/')[-1] ))

    name = r.html.search('<title data-th="server">{}</title>')[0]
    if name is None:
        print('No matching name')

    sku = r.html.search('","storeSkuNumber":"{}","upcGtin13":')[0]
    if sku is None:
        print('No matching SKU')

    upc = r.html.search('"upc":"{}","productType":')[0]
    if upc is None:
        print('No matching UPC')

    dollars = r.html.search('<span class="price-format__large-symbols">$</span><span>{}</span>')[0]
    cents = r.html.search('</span><span class="price-format__large-symbols">{}</span></div></div>')[0]
    if dollars is None or cents is None:
        print('No matching price')
    price = "{}.{}".format(dollars, cents)

    wb = openpyxl.load_workbook("test.xlsx")
    ws = wb.active
    i = 1
    val = str(i)
    cell_val = ws['A' + val].value
    while cell_val != '':
        cell_val = ws['A' + val].value
        i += 1
    i = str(i)
    pallet = input("What pallet number is this item from(what folder should it be added to)?\n")
    ws['A' + i] = pallet
    ws['B' + i] = name
    ifBox = input("Is the item in box? Answer 'y' for yes and 'n' for no.\n")
    if ifBox == 'y':
        ws['C' + i] = "in box"
    elif ifBox == 'n':
        ws['C' + i] = "no box"
    ws['E' + i] = upc
    ws['F' + i] = sku
    ws['G' + i] = price
    wb.save("test.xlsx")

def main():
    inp = input("Enter a UPC code or anything that should work in HomeDepot.com search bar\n")
    get_info("https://www.homedepot.com/s/{}?".format( inp ))

if __name__ == "__main__":
    main()
